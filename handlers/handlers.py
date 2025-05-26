from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton, FSInputFile
from aiogram.filters import Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext

from database.database import SessionLocal, Base, engine
from models.models import User, Task, Report
from utils.utils import is_admin, send_task_notification

from datetime import datetime
import matplotlib.pyplot as plt
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pytz

# Бишкек убакыт зонасы
kz_tz = pytz.timezone("Asia/Bishkek")

# Инициализация
dp = Dispatcher()
Base.metadata.create_all(bind=engine)

# Состояния FSM
class AssignTask(StatesGroup):
    waiting_for_user_id = State()
    waiting_for_task_text = State()

class DoneTask(StatesGroup):
    waiting_for_task_id = State()

# Старт
@dp.message(Command("start"))
async def start_handler(message: Message):
    session = SessionLocal()
    try:
        user = session.query(User).filter(User.user_id == message.from_user.id).first()
        if not user:
            is_admin_flag = True if message.from_user.id == 757804536 else False
            full_name = f"{message.from_user.first_name or ''} {message.from_user.last_name or ''}".strip()
            user = User(
                user_id=message.from_user.id,
                username=message.from_user.username,
                full_name=full_name,
                is_admin=is_admin_flag
            )
            session.add(user)
            session.commit()
        await message.answer("Привет! Я бот для управления задачами.\nВоспользуйтесь командой /menu.")
    finally:
        session.close()


# Меню
@dp.message(Command("menu"))
async def show_menu(message: Message):
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📌 Назначить задание (/assign)")],
            [KeyboardButton(text="✅ Завершить задание (/done)")],
            [KeyboardButton(text="🗒 Мои задания (/mytasks)"), KeyboardButton(text="📊 Отчёт (/report)")],
            [KeyboardButton(text="👥 Пользователи (/users)")]
        ],
        resize_keyboard=True
    )
    await message.answer("📋 Главное меню:", reply_markup=keyboard)

# Назначение задания
@dp.message(F.text == "📌 Назначить задание (/assign)")
@dp.message(Command("assign"))
async def assign_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда доступна только администраторам.")
        return
    await message.answer("👤 Введите user_id пользователя, которому хотите назначить задание:")
    await state.set_state(AssignTask.waiting_for_user_id)

@dp.message(AssignTask.waiting_for_user_id)
async def assign_user_id_input(message: Message, state: FSMContext):
    try:
        user_id = int(message.text)
        await state.update_data(user_id=user_id)
        await message.answer("✏️ Введите текст задания:")
        await state.set_state(AssignTask.waiting_for_task_text)
    except ValueError:
        await message.answer("❌ Введите корректный user_id (только число).")

@dp.message(AssignTask.waiting_for_task_text)
async def assign_task_text_input(message: Message, state: FSMContext):
    data = await state.get_data()
    user_id = data['user_id']
    task_text = message.text
    file_id = message.document.file_id if message.document else None

    session = SessionLocal()
    try:
        user = session.query(User).filter(User.user_id == user_id).first()
        if not user:
            await message.answer("❌ Пользователь с таким user_id не найден.")
            return

        task = Task(description=task_text, assigned_to_user_id=user_id)
        if file_id:
            task.document_file_id = file_id

        session.add(task)
        session.commit()

        if file_id:
            await message.bot.send_document(
                user_id,
                file_id,
                caption=f"🆕 Вам назначено новое задание (ID: {task.task_id}):\n{task_text}"
            )
        else:
            await send_task_notification(message.bot, user_id, task)

        await message.answer(f"✅ Задание назначено: user_id {user_id} (ID: {task.task_id})")
    finally:
        session.close()
    await state.clear()

# Завершение задания
@dp.message(F.text == "✅ Завершить задание (/done)")
@dp.message(Command("done"))
async def done_start(message: Message, state: FSMContext):
    await message.answer("🔢 Введите ID выполненного задания (task_id):")
    await state.set_state(DoneTask.waiting_for_task_id)

@dp.message(DoneTask.waiting_for_task_id)
async def done_get_task_id(message: Message, state: FSMContext):
    try:
        task_id = int(message.text)
        session = SessionLocal()
        try:
            task = session.query(Task).filter(
                Task.task_id == task_id,
                Task.assigned_to_user_id == message.from_user.id
            ).first()

            if not task:
                await message.answer("❌ Задание не найдено или не относится к вам.")
                return

            # Бишкек убактысы
            task.status = 'done'
            task.done_at = datetime.now(kz_tz)

            # ✅ full_name генерациясы
            full_name = f"{message.from_user.first_name or ''} {message.from_user.last_name or ''}".strip()

            report = Report(
                task_id=task.task_id,
                user_id=message.from_user.id,
                report_text=f"Пользователь {full_name} завершил задание.",
                created_at=datetime.now(kz_tz)
            )
            session.add(report)
            session.commit()

            await message.answer(f"✅ Задание {task_id} помечено как выполненное.")
        finally:
            session.close()
        await state.clear()
    except ValueError:
        await message.answer("❗ Введите корректный task_id (число).")

# Мои задания
@dp.message(F.text == "🗒 Мои задания (/mytasks)")
@dp.message(Command("mytasks"))
async def mytasks_handler(message: Message):
    session = SessionLocal()
    try:
        tasks = session.query(Task).filter(Task.assigned_to_user_id == message.from_user.id).all()
        if not tasks:
            await message.answer("📭 У вас нет заданий.")
        else:
            text = "📋 Ваши задания:\n"
            for t in tasks:
                text += f"ID: {t.task_id}, Статус: {t.status}, Текст: {t.description}\n"
            await message.answer(text)
    finally:
        session.close()

# Отчет
@dp.message(F.text == "📊 Отчёт (/report)")
@dp.message(Command("report"))
async def report_handler(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для админов.")
        return

    session = SessionLocal()
    try:
        tasks = session.query(Task).all()
        if not tasks:
            await message.answer("❌ Заданий не найдено.")
            return

        done_tasks = [t for t in tasks if t.status == 'done' and t.done_at]
        dates = [t.done_at.astimezone(kz_tz).date() for t in done_tasks]
        counter = Counter(dates)
        dates_sorted = sorted(counter.keys())
        counts = [counter[d] for d in dates_sorted]

        if counts:
            plt.figure(figsize=(8, 4))
            plt.bar([d.strftime("%Y-%m-%d") for d in dates_sorted], counts)
            plt.title("Выполненные задачи по датам")
            plt.xlabel("Дата")
            plt.ylabel("Количество")
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig("report.png")
            plt.close()
        else:
            plt.figure(figsize=(6, 3))
            plt.text(0.5, 0.5, "Нет завершённых заданий", fontsize=14, ha='center')
            plt.axis('off')
            plt.savefig("report.png")
            plt.close()

        # Excel отчет
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Отчёты"

        headers = [
            "ID задания", "Имя пользователя", "ID пользователя", "Текст задания",
            "Статус", "Дата выполнения", "Дата назначения"
        ]
        sheet.append(headers)

        for task in tasks:
            done_at = (
                task.done_at.astimezone(kz_tz).strftime("%Y-%m-%d %H:%M:%S")
                if task.done_at else ""
            )
            created_at = (
                task.created_at.astimezone(kz_tz).strftime("%Y-%m-%d %H:%M:%S")
                if task.created_at else ""
            )

            user = session.query(User).filter(User.user_id == task.assigned_to_user_id).first()
            username = f"@{user.username}" if user and user.username else user.full_name if user and user.full_name else "Неизвестно"

            sheet.append([
                task.task_id,
                task.assigned_to_user_id or "",
                username,
                task.description or "",
                "завершено" if task.status == 'done' else "не завершено",
                done_at,
                created_at
            ])

        for col in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max_length + 2

        excel_path = "reports.xlsx"
        wb.save(excel_path)

        await message.bot.send_photo(
            message.chat.id,
            FSInputFile("report.png"),
            caption="📊 График выполненных заданий"
        )
        await message.bot.send_document(
            message.chat.id,
            FSInputFile(excel_path),
            caption="📁 Полный Excel-отчёт по всем заданиям"
        )

    except Exception as e:
        await message.answer(f"❌ Ошибка при создании отчета: {str(e)}")
    finally:
        session.close()

# Пользователи
@dp.message(F.text == "👥 Пользователи (/users)")
@dp.message(Command("users"))
async def users_handler(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Эта команда только для админов.")
        return

    session = SessionLocal()
    try:
        users = session.query(User).all()
        if not users:
            await message.answer("❌ Пользователи не найдены.")
        else:
            text = "👥 Список пользователей:\n"
            for u in users:
                name = f"@{u.username}" if u.username else "(без username)"
                text += f"ID: {u.user_id}, Username: {name}, Админ: {u.is_admin}\n"
            await message.answer(text)
    finally:
        session.close()
